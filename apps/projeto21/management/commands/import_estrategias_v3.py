# app/projeto21/management/commands/import_estrategias_v3.py
import csv
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.projeto21.models import Area, Estrategia

# Mapeamentos oficiais das áreas
AREA_MAP = {
    "Família": "Fa",
    "Igreja": "Ig",
    "Escola": "Es",
    "Amigos": "Am",
    "Comunidade": "Co",
    "Eu mesmo": "Eu",
}

DIMENSOES_VALIDAS = {"Básico", "Desafio", "Avançado"}

ID_SIG_RE = re.compile(r"^[A-Z][a-z][BDA]\d{3}-[13]\w\d{2}[sd]$")

def norm_freq_code(freq_text: str, periodo_text: str) -> tuple[str, str]:
    """Normaliza frequência/período em códigos curtos."""
    f = (freq_text or "").strip().lower()
    p = (periodo_text or "").strip().lower()

    # Período -> código
    if "semana" in p:
        # "3 semanas", "1 semana", "2 semanas"
        n = int(p.split()[0])
        pcode = f"{n:02d}s"
    elif "dia" in p:
        n = int(p.split()[0])
        pcode = f"{n:02d}d"
    else:
        pcode = "00d"

    # Frequência -> código
    if f.startswith("diário"):
        fcode = "1d"
    elif f.startswith("dia sim"):   # Dia sim-dia não
        fcode, pcode = "3s", "03s"
    elif f.startswith("3x semana") or f.startswith("3x na semana"):
        fcode = "3s"
    elif f.startswith("semanal 2–3x") or f.startswith("semanal (2–3x)"):
        fcode = "3s"
    elif f.startswith("semanal"):
        fcode = "1s"
    elif f.startswith("programa"):
        # assumimos 1x/dia durante o programa
        fcode = "1d"
    else:
        # fallback seguro
        fcode = "1d"

    return fcode, pcode


class Command(BaseCommand):
    help = "Importa estratégias do Projeto21 (XLSX/CSV) para o BD com upsert por id_sig."

    def add_arguments(self, parser):
        parser.add_argument("--file", "-f", required=True, help="Caminho do arquivo XLSX ou CSV.")
        parser.add_argument("--sheet", "-s", default=None, help="Nome da planilha (XLSX).")
        parser.add_argument("--dry-run", action="store_true", help="Valida sem gravar no banco.")
        parser.add_argument("--verbose", action="store_true", help="Exibe detalhes linha a linha.")
        parser.add_argument("--upsert-by",
                            default="id_sig",
                            choices=["id_sig", "chaves"],
                            help='Chave de upsert: "id_sig" (padrão) ou "chaves" (area+dimensao+ordem).')
        parser.add_argument("--limit", type=int, default=0, help="Importar apenas N linhas (debug).")

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"Arquivo não encontrado: {path}")

        # Leitura
        if path.suffix.lower() == ".csv":
            rows = self._read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            rows = self._read_xlsx(path, options.get("sheet"))
        else:
            raise CommandError("Formato não suportado. Use .csv ou .xlsx/.xlsm/.xls")

        if options.get("limit"):
            rows = rows[: options["limit"]]

        created, updated, skipped = 0, 0, 0

        # Mapa de áreas cacheado
        area_cache: dict[str, Area] = {}

        @transaction.atomic
        def do_import():
            nonlocal created, updated, skipped

            for i, r in enumerate(rows, start=1):
                try:
                    id_sig = (r.get("ID") or "").strip()
                    area_nome = (r.get("Area") or "").strip()
                    dimensao = (r.get("Dimensao") or "").strip()
                    estrategia_txt = (r.get("Estrategia") or "").strip()
                    frequencia = (r.get("Frequencia") or "").strip()
                    periodo = (r.get("Periodo") or "").strip()
                    freq_codigo = (r.get("FreqCodigo") or "").strip()
                    periodo_codigo = (r.get("PeriodoCodigo") or "").strip()
                    peso = int(r.get("Peso") or 1)
                    ordem_area = int(r.get("OrdemArea") or 0)
                    ordem_dim = int(r.get("OrdemDimensao") or 0)
                    ordem_est = int(r.get("OrdemEstrategia") or 0)

                    # Validações básicas
                    if area_nome not in AREA_MAP:
                        raise ValueError(f"[linha {i}] Área inválida: {area_nome}")

                    if dimensao not in DIMENSOES_VALIDAS:
                        raise ValueError(f"[linha {i}] Dimensão inválida: {dimensao}")

                    # Garante Area
                    if area_nome not in area_cache:
                        area_obj, _ = Area.objects.get_or_create(
                            nome=area_nome, defaults={"codigo": AREA_MAP[area_nome]}
                        )
                        # Ajusta código se não estiver correto (idempotente)
                        if area_obj.codigo != AREA_MAP[area_nome]:
                            area_obj.codigo = AREA_MAP[area_nome]
                            area_obj.save(update_fields=["codigo"])
                        area_cache[area_nome] = area_obj
                    else:
                        area_obj = area_cache[area_nome]

                    # Normaliza códigos se vazios
                    if not freq_codigo or not periodo_codigo:
                        freq_codigo, periodo_codigo = norm_freq_code(frequencia, periodo)

                    # Valida/gera id_sig se vazio (não é o ideal, mas permite reconstruir)
                    if not id_sig:
                        # Reconstrói com base nos campos; sequencial é desconhecido,
                        # então usa ordem_estrategia como fallback (3 dígitos).
                        seq = f"{ordem_est:03d}"
                        area2 = AREA_MAP[area_nome]
                        nivel = {"Básico": "B", "Desafio": "D", "Avançado": "A"}[dimensao]
                        id_sig = f"{area2}{nivel}{seq}-{freq_codigo}{periodo_codigo}"

                    if not ID_SIG_RE.match(id_sig):
                        raise ValueError(f"[linha {i}] ID inválido: {id_sig}")

                    # UPSERT
                    if options["upsert_by"] == "id_sig":
                        obj, created_flag = Estrategia.objects.update_or_create(
                            id_sig=id_sig,
                            defaults=dict(
                                area=area_obj,
                                dimensao=dimensao,
                                estrategia=estrategia_txt,
                                frequencia=frequencia,
                                periodo=periodo,
                                freq_codigo=freq_codigo,
                                periodo_codigo=periodo_codigo,
                                peso=peso,
                                ordem_area=ordem_area,
                                ordem_dimensao=ordem_dim,
                                ordem_estrategia=ordem_est,
                                ativo=True,
                            ),
                        )
                    else:
                        obj, created_flag = Estrategia.objects.update_or_create(
                            area=area_obj,
                            dimensao=dimensao,
                            ordem_estrategia=ordem_est,
                            defaults=dict(
                                id_sig=id_sig,
                                estrategia=estrategia_txt,
                                frequencia=frequencia,
                                periodo=periodo,
                                freq_codigo=freq_codigo,
                                periodo_codigo=periodo_codigo,
                                peso=peso,
                                ordem_area=ordem_area,
                                ordem_dimensao=ordem_dim,
                                ativo=True,
                            ),
                        )

                    if created_flag:
                        created += 1
                        if options["verbose"]:
                            self.stdout.write(self.style.SUCCESS(f"[{i}] criado: {obj.id_sig}"))
                    else:
                        updated += 1
                        if options["verbose"]:
                            self.stdout.write(self.style.WARNING(f"[{i}] atualizado: {obj.id_sig}"))

                except Exception as e:
                    skipped += 1
                    self.stderr.write(self.style.ERROR(f"✖ [linha {i}] {e}"))

            if options["dry_run"]:
                self.stdout.write(self.style.WARNING("Dry-run: efetuando rollback"))
                raise transaction.TransactionManagementError("rollback on dry-run")

        if options["dry_run"]:
            try:
                do_import()
            except transaction.TransactionManagementError:
                pass  # esperado no dry-run
        else:
            do_import()

        self.stdout.write(self.style.SUCCESS(f"Concluído: {created} criados, {updated} atualizados, {skipped} ignorados."))

    # --- helpers de leitura ---
    def _read_csv(self, path: Path):
        with path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)

    def _read_xlsx(self, path: Path, sheet_name: str | None):
        try:
            import openpyxl  # noqa: F401
        except ImportError as e:
            raise CommandError("Instale openpyxl para leitura de .xlsx: pip install openpyxl") from e

        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            # Converte numéricos para string onde aplicável
            for k in ("Peso", "OrdemArea", "OrdemDimensao", "OrdemEstrategia"):
                if item.get(k) is not None:
                    item[k] = str(item[k])
            rows.append(item)
        return rows
